import os
import sqlite3
import openpyxl
from contextlib import closing
from typing import List, Dict, Optional

def get_default_db_path() -> str:
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    return os.path.join(base_dir, 'productos.db')

def inicializar_bd(excel_path: str = 'productos_coronel', db_path: Optional[str] = None) -> Optional[str]:
    """Convierte el Excel más reciente a SQLite para consultas rápidas"""
    try:
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        excel_dir = os.path.join(base_dir, excel_path)
        if not db_path:
            db_path = get_default_db_path()
            
        # Buscar Excel más reciente
        archivos = [f for f in os.listdir(excel_dir) if f.lower().endswith('.xlsx')]
        if not archivos:
            return None
            
        archivo_reciente = max(archivos, key=lambda f: os.path.getmtime(os.path.join(excel_dir, f)))
        ruta_excel = os.path.join(excel_dir, archivo_reciente)
        
        with closing(sqlite3.connect(db_path)) as conn:
            cursor = conn.cursor()
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS productos (
                    codigo TEXT PRIMARY KEY,
                    codigo_barra TEXT,
                    descripcion TEXT,
                    precio TEXT,
                    imagen_url TEXT,
                    imagen_local TEXT,
                    variante TEXT,
                    categoria TEXT,
                    subcategoria TEXT,
                    peso_kg TEXT,
                    ancho_cm TEXT,
                    alto_cm TEXT,
                    profundidad_cm TEXT
                )
            ''')
            
            # Cargar datos desde Excel
            wb = openpyxl.load_workbook(ruta_excel)
            ws = wb.active
            
            # Limpiar datos existentes
            cursor.execute("DELETE FROM productos")
            
            # Insertar nuevos datos (solo código y código de barras del Excel)
            for row in range(2, ws.max_row + 1):
                codigo = str(ws.cell(row=row, column=1).value).strip().replace(" ", "")  # Columna A: Código
                codigo_barra = str(ws.cell(row=row, column=3).value).strip()  # Columna C: C.Barra
                
                if codigo:
                    cursor.execute(
                        "INSERT OR REPLACE INTO productos (codigo, codigo_barra) VALUES (?, ?)",
                        (codigo, codigo_barra)
                    )
            
            conn.commit()
            return db_path
        
    except Exception as e:
        print(f"Error inicializando SQLite: {str(e)}")
        return None

def obtener_codigo_barra(code: str, db_path: Optional[str] = None) -> str:
    """Consulta rápida a SQLite para obtener el código de barras"""
    if not db_path:
        db_path = get_default_db_path()
    try:
        with closing(sqlite3.connect(db_path)) as conn:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT codigo_barra FROM productos WHERE codigo = ?",
                (code.strip(),)
            )
            result = cursor.fetchone()
            return result[0] if result else ""
    except Exception as e:
        print(f"Error consultando SQLite para código de barras: {str(e)}")
        return ""

def consolidar_productos(todos_los_productos: List[Dict], db_path: Optional[str] = None):
    """Guarda todos los datos de productos en la base de datos SQLite"""
    if not db_path:
        db_path = get_default_db_path()
        
    try:
        with closing(sqlite3.connect(db_path)) as conn:
            cursor = conn.cursor()
            
            for producto in todos_los_productos:
                codigo = producto.get('codigo', '')
                codigo_barra = producto.get('codigo_de_barras', '')
                
                # Asegurar que el registro existe
                cursor.execute("INSERT OR IGNORE INTO productos (codigo, codigo_barra) VALUES (?, ?)", (codigo, codigo_barra))
                
                # Actualizar los campos obtenidos del scraping, preservando dimensiones preexistentes
                cursor.execute("""
                    UPDATE productos SET
                        descripcion = ?,
                        precio = ?,
                        imagen_url = ?,
                        imagen_local = ?,
                        variante = ?,
                        categoria = ?,
                        subcategoria = ?
                    WHERE codigo = ?
                """, (
                    producto.get('descripcion', ''),
                    producto.get('precio', ''),
                    producto.get('imagen_url', ''),
                    producto.get('imagen_local', ''),
                    producto.get('variante', ''),
                    producto.get('categoria', ''),
                    producto.get('subcategoria', ''),
                    codigo
                ))
            
            # Eliminar productos con descripción vacía
            cursor.execute("DELETE FROM productos WHERE descripcion = '' OR descripcion IS NULL")
            conn.commit()
            print(f"Se consolidaron {len(todos_los_productos)} productos en la base de datos")
            return todos_los_productos
            
    except Exception as e:
        print(f"Error consolidando productos en SQLite: {str(e)}")
        return False

def obtener_productos_de_db(db_path: Optional[str] = None) -> List[Dict]:
    """Obtiene todos los productos de la base de datos en formato lista de diccionarios"""
    if not db_path:
        db_path = get_default_db_path()
        
    try:
        with closing(sqlite3.connect(db_path)) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM productos")
            productos = cursor.fetchall()
            
            lista_productos = []
            for producto in productos:
                lista_productos.append({
                    'codigo': producto[0],
                    'codigo_de_barras': producto[1],
                    'descripcion': producto[2],
                    'precio': producto[3],
                    'imagen_url': producto[4],
                    'imagen_local': producto[5],
                    'variante': producto[6].replace('-', '') if producto[6] else None,
                    'categoria': producto[7],
                    'subcategoria': producto[8],
                    'peso_kg': producto[9],
                    'ancho_cm': producto[10],
                    'alto_cm': producto[11],
                    'profundidad_cm': producto[12]
                })
            return lista_productos
    except Exception as e:
        print(f"Error leyendo productos de SQLite: {str(e)}")
        return []

def actualizar_dimensiones_en_bd(productos: List[Dict], dimensiones: List[Dict], db_path: Optional[str] = None):
    """Actualiza la base de datos SQLite con las dimensiones de envío"""
    if not db_path:
        db_path = get_default_db_path()
        
    try:
        with closing(sqlite3.connect(db_path)) as conn:
            cursor = conn.cursor()
            for producto, dim in zip(productos, dimensiones):
                cursor.execute("""
                    UPDATE productos SET
                        peso_kg = ?,
                        ancho_cm = ?,
                        alto_cm = ?,
                        profundidad_cm = ?
                    WHERE codigo = ?
                """, (
                    dim.get('peso_kg', ''),
                    dim.get('ancho_cm', ''),
                    dim.get('alto_cm', ''),
                    dim.get('profundidad_cm', ''),
                    producto['codigo']
                ))
            conn.commit()
            print(f"Actualizadas dimensiones para {len(dimensiones)} productos")
    except Exception as e:
        print(f"Error actualizando dimensiones en SQLite: {str(e)}")
