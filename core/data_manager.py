import sqlite3
import os
import openpyxl
import logging
from contextlib import closing
from typing import List, Dict, Any, Optional

from config import settings # Usar las rutas y configuraciones centralizadas
from colorama import Fore, Style # Mantener colorama por ahora para los prints, aunque se podría pasar a logging

logger = logging.getLogger(__name__)

# Definición de la estructura de la tabla de productos
# Esta podría moverse a un modelo si se usa un ORM simple o para mayor claridad.
TABLE_NAME = "productos"
TABLE_COLUMNS = [
    "codigo TEXT PRIMARY KEY",
    "codigo_barra TEXT",
    "descripcion TEXT",
    "precio TEXT", # Considerar almacenar como NUMERIC o INTEGER (en centavos) si se hacen cálculos
    "imagen_url TEXT",
    "imagen_local TEXT", # Ruta relativa a IMG_SCRAPING_DIR
    "variante TEXT",
    "categoria TEXT",
    "subcategoria TEXT",
    "peso_kg TEXT", # Considerar NUMERIC
    "ancho_cm TEXT", # Considerar NUMERIC
    "alto_cm TEXT", # Considerar NUMERIC
    "profundidad_cm TEXT" # Considerar NUMERIC
]

def _get_db_connection() -> sqlite3.Connection:
    """Establece y devuelve una conexión a la base de datos."""
    try:
        # Asegurarse que el directorio de datos exista
        os.makedirs(settings.DATA_DIR, exist_ok=True)
        conn = sqlite3.connect(settings.DB_PATH, timeout=10)
        logger.debug(f"Conexión a la base de datos {settings.DB_PATH} establecida.")
        return conn
    except sqlite3.Error as e:
        logger.error(f"Error al conectar con la base de datos {settings.DB_PATH}: {e}", exc_info=True)
        raise # Relanzar para que el llamador maneje el error de conexión

def _create_products_table_if_not_exists(conn: sqlite3.Connection):
    """Crea la tabla de productos si no existe."""
    try:
        with closing(conn.cursor()) as cursor:
            columns_sql = ",\n    ".join(TABLE_COLUMNS)
            create_table_sql = f"CREATE TABLE IF NOT EXISTS {TABLE_NAME} (\n    {columns_sql}\n);"
            cursor.execute(create_table_sql)
            conn.commit()
            logger.info(f"Tabla '{TABLE_NAME}' verificada/creada exitosamente.")
    except sqlite3.Error as e:
        logger.error(f"Error al crear la tabla '{TABLE_NAME}': {e}", exc_info=True)
        raise

def inicializar_base_de_datos_desde_excel() -> bool:
    """
    Busca el archivo Excel de precios más reciente en PRODUCTOS_CORONEL_DIR,
    y (re)crea/actualiza la tabla de productos en la base de datos SQLite
    con los códigos y códigos de barras del Excel.
    Retorna True si fue exitoso, False en caso contrario.
    """
    logger.info("Iniciando proceso de inicialización/actualización de BD desde Excel.")
    try:
        excel_dir = settings.PRODUCTOS_CORONEL_DIR
        archivos_excel = [f for f in os.listdir(excel_dir) if f.lower().endswith('.xlsx') and not f.startswith('~$')]
        if not archivos_excel:
            logger.warning(f"No se encontraron archivos Excel en {excel_dir}.")
            # Considerar si crear la tabla vacía es el comportamiento deseado o si debe fallar.
            # Por ahora, crearemos la tabla vacía si no existe.
            with _get_db_connection() as conn:
                _create_products_table_if_not_exists(conn)
            return True # O False si se considera un fallo no encontrar Excel

        archivo_excel_reciente = max(archivos_excel, key=lambda f: os.path.getmtime(os.path.join(excel_dir, f)))
        ruta_excel_completa = os.path.join(excel_dir, archivo_excel_reciente)
        logger.info(f"Usando el archivo Excel más reciente: {ruta_excel_completa}")

        wb = openpyxl.load_workbook(ruta_excel_completa, data_only=True) # data_only para obtener valores de fórmulas
        ws = wb.active

        productos_excel = []
        # Asumir que la primera fila es de encabezados, los datos empiezan en la fila 2.
        # Columna A (1): Código, Columna C (3): C.Barra
        for row_idx in range(2, ws.max_row + 1):
            codigo_val = ws.cell(row=row_idx, column=1).value
            cod_barra_val = ws.cell(row=row_idx, column=3).value

            codigo = str(codigo_val).strip().replace(" ", "") if codigo_val else None
            codigo_barra = str(cod_barra_val).strip() if cod_barra_val else None

            if codigo: # Solo procesar si hay un código de producto
                productos_excel.append({'codigo': codigo, 'codigo_barra': codigo_barra})

        if not productos_excel:
            logger.warning(f"No se extrajeron datos de productos del archivo Excel: {archivo_excel_reciente}")
            # De nuevo, decidir si esto es un fallo o no.
            with _get_db_connection() as conn:
                _create_products_table_if_not_exists(conn)
            return True


        with _get_db_connection() as conn:
            _create_products_table_if_not_exists(conn) # Asegurar que la tabla exista

            with closing(conn.cursor()) as cursor:
                # Estrategia: Actualizar códigos de barra para códigos existentes, insertar nuevos.
                # Podríamos borrar todo y reinsertar si esa es la lógica deseada (como en el original).
                # El original hacía "DELETE FROM productos" y luego "INSERT OR REPLACE".
                # Esto es similar a borrar y reinsertar solo los campos codigo y codigo_barra.

                logger.info(f"Borrando datos existentes de la tabla '{TABLE_NAME}' para recargar desde Excel (solo códigos y códigos de barras).")
                cursor.execute(f"DELETE FROM {TABLE_NAME}")
                # Si solo se quieren actualizar códigos de barras y añadir nuevos productos del Excel:
                # No borrar, y usar INSERT OR IGNORE para nuevos, y UPDATE para existentes.
                # Pero para replicar el original que limpia y recarga:

                insert_sql = f"INSERT INTO {TABLE_NAME} (codigo, codigo_barra) VALUES (?, ?)"
                cursor.executemany(insert_sql, [(p['codigo'], p['codigo_barra']) for p in productos_excel])
                conn.commit()
                logger.info(f"{cursor.rowcount} productos (códigos y códigos de barra) insertados/actualizados en '{TABLE_NAME}' desde Excel.")
        return True

    except FileNotFoundError:
        logger.error(f"El directorio de Excel {settings.PRODUCTOS_CORONEL_DIR} no fue encontrado.", exc_info=True)
        return False
    except Exception as e:
        logger.error(f"Error al inicializar la base de datos desde Excel: {e}", exc_info=True)
        return False

def guardar_o_actualizar_producto(producto_data: Dict[str, Any]) -> bool:
    """
    Guarda un nuevo producto o actualiza uno existente en la base de datos.
    Usa el 'codigo' del producto como clave primaria.
    """
    if not producto_data.get('codigo'):
        logger.warning("Intento de guardar producto sin 'codigo'. Operación abortada.")
        return False

    codigo_producto = producto_data['codigo']
    logger.debug(f"Guardando/Actualizando producto con código: {codigo_producto}")

    try:
        with _get_db_connection() as conn:
            _create_products_table_if_not_exists(conn) # Asegurar que la tabla exista
            with closing(conn.cursor()) as cursor:
                # Construir la sentencia INSERT OR REPLACE dinámicamente
                # para manejar solo las columnas presentes en producto_data
                column_names = []
                column_values = []
                for key, value in producto_data.items():
                    # Validar que la columna exista en nuestra definición (TABLE_COLUMNS)
                    # Esto es una salvaguarda, asume que las claves de producto_data son nombres de columnas.
                    if any(key == col.split()[0] for col in TABLE_COLUMNS):
                        column_names.append(key)
                        column_values.append(value)
                    else:
                        logger.warning(f"Clave '{key}' en producto_data no es una columna válida en la tabla '{TABLE_NAME}'. Se omitirá.")

                if not column_names:
                    logger.warning(f"No hay columnas válidas para guardar para el producto {codigo_producto}.")
                    return False

                placeholders = ", ".join(["?"] * len(column_names))
                columns_sql_segment = ", ".join(column_names)

                # Usar INSERT OR REPLACE para actualizar si existe, o insertar si es nuevo.
                sql = f"INSERT OR REPLACE INTO {TABLE_NAME} ({columns_sql_segment}) VALUES ({placeholders})"

                cursor.execute(sql, tuple(column_values))
                conn.commit()
                logger.info(f"Producto '{codigo_producto}' guardado/actualizado exitosamente ({cursor.rowcount} fila afectada).")
                return True
    except sqlite3.Error as e:
        logger.error(f"Error de base de datos al guardar/actualizar producto '{codigo_producto}': {e}", exc_info=True)
        return False
    except Exception as e:
        logger.error(f"Error inesperado al guardar/actualizar producto '{codigo_producto}': {e}", exc_info=True)
        return False

def guardar_multiples_productos(lista_productos: List[Dict[str, Any]], sobrescribir_descripcion_vacia=True) -> int:
    """
    Guarda o actualiza una lista de productos en la base de datos.
    Retorna el número de productos procesados exitosamente.
    Si sobrescribir_descripcion_vacia es True, los productos cuya descripción esté vacía
    en la lista_productos NO se insertarán/actualizarán si ya existen con descripción.
    Si es False, se actualizarán incluso si la nueva descripción está vacía.
    (Esta lógica de descripción vacía es del original `consolidar_todo_en_base_de_datos`)
    """
    # La lógica original de `consolidar_todo_en_base_de_datos` tenía una parte que
    # eliminaba productos con descripción vacía DESPUÉS de insertarlos/reemplazarlos.
    # Y otra parte que para códigos con guion, siempre hacía INSERT, y sin guion, INSERT OR REPLACE.
    # Esto se simplificará aquí a un INSERT OR REPLACE para todos, y la limpieza de descripciones vacías
    # se puede hacer en una función separada o como parte de este proceso si se requiere.

    # Por ahora, cada producto se maneja con guardar_o_actualizar_producto.
    # La lógica de "no insertar si la descripción está vacía" es un poco más compleja
    # si se quiere evitar el reemplazo de una descripción existente con una vacía.

    # Replicando la lógica original de `consolidar_todo_en_base_de_datos` más de cerca:
    # 1. Insertar o reemplazar todos los productos.
    # 2. Eliminar productos donde la descripción final sea vacía.

    productos_exitosos = 0
    logger.info(f"Iniciando guardado de {len(lista_productos)} productos.")

    try:
        with _get_db_connection() as conn:
            _create_products_table_if_not_exists(conn) # Asegurar tabla

            for producto_data in lista_productos:
                if guardar_o_actualizar_producto(producto_data): # Usa la conexión implícita de la función
                    productos_exitosos += 1

            # Paso 2 del original: Eliminar productos con descripción vacía
            if sobrescribir_descripcion_vacia: # Esta condición es confusa respecto al nombre
                                            # El original siempre borraba si la descripción era vacía después de la inserción/reemplazo.
                                            # Si `sobrescribir_descripcion_vacia` es True, significa que una desc vacía
                                            # PUEDE haber sobrescrito una no vacía. Si es False, una desc vacía NO debería
                                            # haber sobrescrito una no vacía.
                                            # La limpieza posterior es independiente de esto.
                pass # No hacer nada aquí si la lógica de guardar_o_actualizar ya maneja esto.

            # Limpieza final de productos con descripción vacía (como en el original)
            with closing(conn.cursor()) as cursor:
                cursor.execute(f"DELETE FROM {TABLE_NAME} WHERE descripcion = '' OR descripcion IS NULL")
                conn.commit()
                deleted_count = cursor.rowcount
                if deleted_count > 0:
                    logger.info(f"Se eliminaron {deleted_count} productos con descripción vacía de la base de datos.")

    except Exception as e:
        logger.error(f"Error durante el guardado de múltiples productos: {e}", exc_info=True)

    logger.info(f"Guardado de múltiples productos finalizado. {productos_exitosos} de {len(lista_productos)} procesados para inserción/actualización.")
    return productos_exitosos


def obtener_producto_por_codigo(codigo: str) -> Optional[Dict[str, Any]]:
    """Obtiene un producto por su código."""
    logger.debug(f"Obteniendo producto con código: {codigo}")
    try:
        with _get_db_connection() as conn:
            conn.row_factory = sqlite3.Row # Para acceder a columnas por nombre
            with closing(conn.cursor()) as cursor:
                cursor.execute(f"SELECT * FROM {TABLE_NAME} WHERE codigo = ?", (codigo,))
                row = cursor.fetchone()
                if row:
                    logger.debug(f"Producto {codigo} encontrado en BD.")
                    return dict(row)
                else:
                    logger.debug(f"Producto {codigo} NO encontrado en BD.")
                    return None
    except sqlite3.Error as e:
        logger.error(f"Error de BD al obtener producto '{codigo}': {e}", exc_info=True)
        return None

def obtener_todos_los_productos() -> List[Dict[str, Any]]:
    """Obtiene todos los productos de la base de datos."""
    logger.info("Obteniendo todos los productos de la base de datos.")
    productos = []
    try:
        with _get_db_connection() as conn:
            _create_products_table_if_not_exists(conn) # Asegurar que la tabla exista
            conn.row_factory = sqlite3.Row
            with closing(conn.cursor()) as cursor:
                cursor.execute(f"SELECT * FROM {TABLE_NAME} ORDER BY codigo") # Ordenar para consistencia
                rows = cursor.fetchall()
                for row in rows:
                    productos.append(dict(row))
        logger.info(f"Se obtuvieron {len(productos)} productos de la base de datos.")
    except sqlite3.Error as e:
        logger.error(f"Error de BD al obtener todos los productos: {e}", exc_info=True)
    return productos

def obtener_codigo_barra_por_codigo_producto(codigo_producto: str) -> Optional[str]:
    """Consulta rápida para obtener el código de barras de un producto."""
    producto = obtener_producto_por_codigo(codigo_producto)
    if producto:
        return producto.get('codigo_barra')
    return None


# --- Funciones de Limpieza (del antiguo limpiar_productos_coronel.py) ---
def limpiar_espacios_en_codigos_producto() -> bool:
    """
    Revisa todos los códigos de producto en la tabla 'productos'.
    Si un código contiene espacios, lo actualiza eliminándolos.
    Esta es una operación destructiva sobre la columna 'codigo'.
    PRECAUCIÓN: Modificar claves primarias puede tener efectos secundarios si hay relaciones.
    En este caso, 'codigo' es PK, así que se debe hacer con cuidado.
    Una estrategia más segura sería crear una nueva tabla o actualizar en una transacción
    manejando duplicados si la limpieza genera códigos que ya existen.

    El original creaba una tabla `productos_limpios`. Aquí se modificará la tabla original
    para simplificar, pero se advierte del riesgo.
    Si se requiere la tabla `productos_limpios`, se puede adaptar.
    """
    logger.info("Iniciando limpieza de espacios en códigos de producto.")
    modificados = 0
    productos_a_actualizar = [] # (nuevo_codigo, descripcion, ..., antiguo_codigo)

    try:
        with _get_db_connection() as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            # Obtener todos los productos
            cursor.execute(f"SELECT * FROM {TABLE_NAME}")
            productos = cursor.fetchall()

            for row_dict in [dict(row) for row in productos]:
                codigo_original = row_dict.get('codigo')
                if codigo_original and ' ' in codigo_original:
                    codigo_limpio = codigo_original.replace(' ', '')
                    # Verificar si el código limpio ya existe (y no es el mismo producto)
                    cursor.execute(f"SELECT codigo FROM {TABLE_NAME} WHERE codigo = ? AND codigo != ?", (codigo_limpio, codigo_original))
                    existente_con_codigo_limpio = cursor.fetchone()

                    if existente_con_codigo_limpio:
                        logger.warning(f"Al limpiar '{codigo_original}' a '{codigo_limpio}', "
                                       f"el código limpio ya existe. Se omitirá la actualización de '{codigo_original}'. "
                                       f"Considere fusionar o eliminar '{codigo_original}' manualmente.")
                        continue # Saltar este producto

                    # Preparar para actualizar: copiar todos los datos al nuevo código y borrar el antiguo
                    # Esto es complejo porque el código es PK. Es más fácil si se puede hacer UPDATE directo.
                    # Si el PK no fuera TEXT y autoincremental, sería UPDATE productos SET codigo = ? WHERE id = ?
                    # Como es PK, la mejor forma es INSERT con el nuevo código y DELETE el viejo.

                    # Crear una nueva entrada con el código limpio y todos los demás datos
                    nueva_data_producto = row_dict.copy()
                    nueva_data_producto['codigo'] = codigo_limpio

                    # Eliminar el producto con el código antiguo
                    cursor.execute(f"DELETE FROM {TABLE_NAME} WHERE codigo = ?", (codigo_original,))

                    # Insertar el producto con el código limpio
                    # Reusar guardar_o_actualizar_producto (pero necesita la conexión directa)
                    column_names = []
                    column_values = []
                    for key, value in nueva_data_producto.items():
                        if any(key == col.split()[0] for col in TABLE_COLUMNS):
                            column_names.append(key)
                            column_values.append(value)

                    if column_names:
                        placeholders = ", ".join(["?"] * len(column_names))
                        columns_sql_segment = ", ".join(column_names)
                        sql_insert = f"INSERT INTO {TABLE_NAME} ({columns_sql_segment}) VALUES ({placeholders})"
                        cursor.execute(sql_insert, tuple(column_values))
                        modificados += 1
                        logger.info(f"Código '{codigo_original}' limpiado a '{codigo_limpio}' y datos transferidos.")
                    else:
                        logger.warning(f"No hay datos válidos para transferir para el código original {codigo_original} después de limpiar.")


            if modificados > 0:
                conn.commit()
                logger.info(f"Limpieza de espacios en códigos completada. {modificados} códigos actualizados.")
            else:
                logger.info("No se encontraron códigos con espacios para limpiar.")

            return True

    except sqlite3.Error as e:
        logger.error(f"Error de BD durante la limpieza de códigos: {e}", exc_info=True)
        return False
    except Exception as e:
        logger.error(f"Error inesperado durante la limpieza de códigos: {e}", exc_info=True)
        return False


if __name__ == '__main__':
    # Bloque de prueba (ejecutar con python -m core.data_manager)
    # Configurar logging para ver salidas
    if not logging.getLogger().hasHandlers():
        from config.logging_config import setup_logging
        setup_logging()

    logger.info("Probando el módulo DataManager...")

    # 1. Inicializar/Actualizar BD desde Excel (asegúrate que haya un Excel en data/productos_coronel/)
    # print(Fore.CYAN + "Intentando inicializar BD desde Excel..." + Style.RESET_ALL)
    # exito_init = inicializar_base_de_datos_desde_excel()
    # print(Fore.GREEN + f"Resultado inicialización: {exito_init}" + Style.RESET_ALL if exito_init else Fore.RED + f"Resultado inicialización: {exito_init}" + Style.RESET_ALL)

    # 2. Guardar/Actualizar un producto de prueba
    # print(Fore.CYAN + "\nIntentando guardar/actualizar producto de prueba..." + Style.RESET_ALL)
    # test_product_1 = {
    #     'codigo': 'TESTPROD001',
    #     'descripcion': 'Producto de Prueba 1',
    #     'precio': '199.99',
    #     'categoria': 'Pruebas',
    #     'codigo_barra': '111222333444',
    #     'imagen_local': 'testprod001.jpg'
    # }
    # exito_save1 = guardar_o_actualizar_producto(test_product_1)
    # print(Fore.GREEN + f"Resultado guardado P1: {exito_save1}" + Style.RESET_ALL if exito_save1 else Fore.RED + f"Resultado guardado P1: {exito_save1}" + Style.RESET_ALL)

    # test_product_2_update = {
    #     'codigo': 'TESTPROD001', # Mismo código
    #     'descripcion': 'Producto de Prueba 1 ACTUALIZADO',
    #     'precio': '210.50',
    #     'subcategoria': 'SubPrueba' # Nueva columna
    # }
    # exito_save2 = guardar_o_actualizar_producto(test_product_2_update)
    # print(Fore.GREEN + f"Resultado guardado P2 (actualización): {exito_save2}" + Style.RESET_ALL if exito_save2 else Fore.RED + f"Resultado guardado P2 (actualización): {exito_save2}" + Style.RESET_ALL)


    # 3. Obtener el producto de prueba
    # print(Fore.CYAN + "\nIntentando obtener producto de prueba..." + Style.RESET_ALL)
    # producto_obtenido = obtener_producto_por_codigo('TESTPROD001')
    # if producto_obtenido:
    #     print(Fore.GREEN + "Producto TESTPROD001 encontrado:" + Style.RESET_ALL)
    #     for k, v in producto_obtenido.items(): print(f"  {k}: {v}")
    # else:
    #     print(Fore.RED + "Producto TESTPROD001 NO encontrado." + Style.RESET_ALL)

    # 4. Obtener todos los productos
    # print(Fore.CYAN + "\nIntentando obtener todos los productos..." + Style.RESET_ALL)
    # todos = obtener_todos_los_productos()
    # print(f"Total de productos en BD: {len(todos)}")
    # if todos:
    #     print("Primeros 2 productos (o menos):")
    #     for p in todos[:2]:
    #         print(f"  Código: {p.get('codigo')}, Desc: {p.get('descripcion')}, Precio: {p.get('precio')}")

    # 5. Guardar múltiples productos
    # print(Fore.CYAN + "\nIntentando guardar múltiples productos..." + Style.RESET_ALL)
    # lista_multi = [
    #     {'codigo': 'MULTI001', 'descripcion': 'Multi Prod 1', 'precio': '10.00', 'categoria': 'Multi'},
    #     {'codigo': 'MULTI002', 'descripcion': '', 'precio': '20.00', 'categoria': 'Multi'}, # Desc vacía
    #     {'codigo': 'MULTI003', 'descripcion': 'Multi Prod 3', 'precio': '30.00', 'categoria': 'Multi'},
    # ]
    # num_guardados = guardar_multiples_productos(lista_multi)
    # print(f"Productos múltiples guardados/actualizados: {num_guardados}")
    # multi002 = obtener_producto_por_codigo('MULTI002') # Debería haberse borrado por desc vacía
    # print(f"MULTI002 después de guardar múltiples: {multi002}")


    # 6. Probar limpieza de códigos (¡CUIDADO! Modifica datos)
    # print(Fore.CYAN + "\nIntentando limpiar espacios en códigos..." + Style.RESET_ALL)
    # Crear datos de prueba para limpieza
    # guardar_o_actualizar_producto({'codigo': 'CODIGO CON ESPACIO', 'descripcion': 'Test Espacio', 'precio': '5.00'})
    # guardar_o_actualizar_producto({'codigo': 'CODIGOOTRO', 'descripcion': 'Test Sin Espacio', 'precio': '6.00'}) # Para probar colisión
    # exito_limpieza = limpiar_espacios_en_codigos_producto()
    # print(Fore.GREEN + f"Resultado limpieza: {exito_limpieza}" + Style.RESET_ALL if exito_limpieza else Fore.RED + f"Resultado limpieza: {exito_limpieza}" + Style.RESET_ALL)
    # p_limpio = obtener_producto_por_codigo('CODIGOCONESPACIO')
    # p_original_con_espacio = obtener_producto_por_codigo('CODIGO CON ESPACIO')
    # print(f"Producto 'CODIGOCONESPACIO' (limpio): {p_limpio}")
    # print(f"Producto 'CODIGO CON ESPACIO' (original): {p_original_con_espacio}")


    logger.info("Pruebas de DataManager finalizadas.")
